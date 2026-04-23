"""
Telegram-бот для роботи з перевізниками (канал + бот). v3.

Нове у v3:
  - Реєстрація перевізника (ПІБ, телефон, ЄДРПОУ, тонаж, тип авто).
    Без реєстрації не можна вводити ціну.
  - Валідація ціни і телефону з поясненнями.
  - Логування у файл + пересилання помилок супер-адміну.
  - Щоденний бекап БД у приватний чат супер-адміна.
  - Кнопка "Обрати переможця" в звіті з авто-нотифікаціями.
  - Редагування оголошення (маршрут, вантаж, вага, дата, контакт).
  - Експорт звіту в Excel (.xlsx).
  - Автозакриття через N годин + нагадування за 2 год.
  - Розсилка всім перевізникам.
  - Blacklist перевізників + історія (перемоги / заявки).
  - Фото вантажу в оголошенні.
  - Фільтри у списку (Усі / 🟢 / 🟡 / 🔴).
  - Статистика (оголошення, середня ціна, топ маршрути/перевізники).
"""
import asyncio
import io
import logging
import os
import re
import traceback
from datetime import datetime, timedelta
from typing import Optional

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from aiogram.filters import Command, CommandObject, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    ErrorEvent,
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaPhoto,
    KeyboardButton,
    Message,
    MessageOriginUser,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)

from database import Database

# ────────────────────  Конфігурація  ────────────────────
BOT_TOKEN = os.getenv("BOT_TOKEN")
SUPER_ADMIN_ID = int(os.getenv("ADMIN_ID", "0"))
CHANNEL_ID = int(os.getenv("CHANNEL_ID", "0"))
BOT_USERNAME = os.getenv("BOT_USERNAME", "").lstrip("@")
DB_PATH = os.getenv("DB_PATH", "bot.db")
AUTO_CLOSE_HOURS = int(os.getenv("AUTO_CLOSE_HOURS", "48"))
REMINDER_BEFORE_HOURS = int(os.getenv("REMINDER_BEFORE_HOURS", "2"))
BACKUP_HOUR_KYIV = int(os.getenv("BACKUP_HOUR_KYIV", "9"))  # 09:00 за Києвом
LOG_FILE = os.getenv("LOG_FILE", "bot.log")

if not BOT_TOKEN or not SUPER_ADMIN_ID or not CHANNEL_ID or not BOT_USERNAME:
    raise RuntimeError(
        "Задайте змінні середовища BOT_TOKEN, ADMIN_ID, CHANNEL_ID, BOT_USERNAME"
    )

# ────────────────────  Логування  ────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
    ],
)
logger = logging.getLogger("pereviznyk-bot")

# ────────────────────  Ініціалізація  ────────────────────
db = Database(DB_PATH)
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())
router = Router()
dp.include_router(router)


# ────────────────────  FSM-стани  ────────────────────
class RegisterStates(StatesGroup):
    full_name = State()
    phone = State()
    edrpou = State()


class RejectReasonStates(StatesGroup):
    waiting_reason = State()


class NewOfferStates(StatesGroup):
    route_from = State()
    route_to = State()
    cargo = State()
    weight = State()
    load_date = State()
    extra_info = State()
    contact_name = State()
    contact_phone = State()
    photo = State()
    confirm = State()


class EditOfferStates(StatesGroup):
    choose_field = State()
    new_value = State()


class PriceInputStates(StatesGroup):
    with_vat = State()
    without_vat = State()


class AddAdminStates(StatesGroup):
    waiting_for_user = State()


class BlacklistStates(StatesGroup):
    reason = State()


class BroadcastStates(StatesGroup):
    text = State()
    confirm = State()


# ────────────────────  Тексти кнопок  ────────────────────
ADMIN_BTN_NEW = "🆕 Нове оголошення"
ADMIN_BTN_LIST = "📋 Активні оголошення"
ADMIN_BTN_ADMINS = "👥 Адміни"
ADMIN_BTN_BROADCAST = "📢 Розсилка"
ADMIN_BTN_STATS = "📊 Статистика"
ADMIN_BTN_PENDING = "📝 Реєстрації"

CARRIER_BTN_MY_PROFILE = "👤 Мій профіль"
CARRIER_BTN_HELP = "❓ Як користуватись"

BTN_WITH_VAT = "📝 Ввести тариф з ПДВ"
BTN_WITHOUT_VAT = "📝 Ввести тариф без ПДВ"
BTN_CONTACT = "📱 Передати контакт"
BTN_CANCEL = "❌ Скасувати"
BTN_SKIP = "⏭ Пропустити"
BTN_CONFIRM = "✅ Опублікувати"
BTN_SHARE_PHONE = "📞 Поділитись номером"


# ────────────────────  Права доступу  ────────────────────
def is_super_admin(user_id: int) -> bool:
    return user_id == SUPER_ADMIN_ID


def is_admin(user_id: int) -> bool:
    return is_super_admin(user_id) or db.is_admin_db(user_id)


# ────────────────────  Утиліти ────────────────────
def html_escape(text: str) -> str:
    if text is None:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def user_profile_link(
    user_id: int, username: Optional[str], first_name: Optional[str]
) -> str:
    """HTML-посилання на профіль Telegram."""
    if username:
        url = f"https://t.me/{username}"
        display = f"@{html_escape(username)}"
    else:
        url = f"tg://user?id={user_id}"
        display = html_escape(first_name or f"id{user_id}")
    return f'<a href="{url}">{display}</a>'


def status_emoji(status: str) -> str:
    return {"open": "🟢", "in_progress": "🟡", "closed": "🔴"}.get(status, "⚪")


def status_text(status: str) -> str:
    return {
        "open": "Відкрито",
        "in_progress": "В роботі",
        "closed": "Закрито",
    }.get(status, status)


def fmt_weight(weight) -> str:
    w = float(weight)
    return f"{w:.0f}" if w.is_integer() else f"{w:g}"


def fmt_price(price) -> str:
    if price is None:
        return "—"
    p = float(price)
    return f"{p:,.0f}".replace(",", " ") if p.is_integer() else f"{p:,.2f}".replace(",", " ")


def channel_post_url(channel_message_id: int) -> str:
    if not channel_message_id:
        return ""
    chan = str(CHANNEL_ID).replace("-100", "", 1)
    return f"https://t.me/c/{chan}/{channel_message_id}"


PRICE_RE = re.compile(r"[^\d.,]")


def parse_price(text: str) -> Optional[float]:
    """Парсер ціни. Повертає None якщо некоректно."""
    if not text:
        return None
    cleaned = PRICE_RE.sub("", text).replace(",", ".")
    if not cleaned:
        return None
    # Якщо кілька крапок — беремо перше число з крапками
    if cleaned.count(".") > 1:
        parts = cleaned.split(".")
        cleaned = parts[0] + "." + "".join(parts[1:])
    try:
        value = float(cleaned)
    except ValueError:
        return None
    if value < 0 or value > 10_000_000:
        return None
    return value


PHONE_RE = re.compile(r"[^\d+]")


def parse_phone(text: str) -> Optional[str]:
    """Нормалізує телефон до +380XXXXXXXXX. Повертає None якщо некоректно."""
    if not text:
        return None
    cleaned = PHONE_RE.sub("", text)
    # Прибрати всі '+' крім першого
    if "+" in cleaned:
        cleaned = "+" + cleaned.replace("+", "")
    digits = cleaned.lstrip("+")
    if len(digits) < 9 or len(digits) > 15:
        return None
    # Український формат: якщо 10 цифр починається з 0 → +38(0...)
    if len(digits) == 10 and digits.startswith("0"):
        return "+38" + digits
    # Якщо 12 цифр починається з 38 → +380...
    if len(digits) == 12 and digits.startswith("380"):
        return "+" + digits
    # Якщо вже з +
    if cleaned.startswith("+"):
        return cleaned
    return "+" + digits


def parse_tonnage(text: str) -> Optional[float]:
    if not text:
        return None
    cleaned = text.replace(",", ".").strip()
    cleaned = re.sub(r"[^\d.]", "", cleaned)
    if not cleaned:
        return None
    try:
        v = float(cleaned)
    except ValueError:
        return None
    if 0 < v <= 100000:
        return v
    return None


def parse_edrpou(text: str) -> Optional[str]:
    """ЄДРПОУ — 8 або 10 цифр, або 10 цифр ІПН (ФОП)."""
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    if len(digits) in (8, 10):
        return digits
    return None


def format_offer_for_channel(offer: dict) -> str:
    status = status_emoji(offer["status"]) + " " + status_text(offer["status"])
    extra = (
        f"\n💬 {html_escape(offer['extra_info'])}"
        if offer.get("extra_info")
        else ""
    )
    date = (
        f"\n📅 Дата завантаження: {html_escape(offer['load_date'])}"
        if offer.get("load_date")
        else ""
    )
    close_at = ""
    if offer.get("auto_close_at"):
        close_at = (
            f"\n⏳ Приймаємо пропозиції до "
            f"{html_escape(offer['auto_close_at'][:16])}"
        )
    return (
        f"<b>Оголошення #{offer['id']}</b> — {status}\n"
        f"━━━━━━━━━━━━━━\n"
        f"🚛 <b>Маршрут:</b> {html_escape(offer['route_from'])} → "
        f"{html_escape(offer['route_to'])}\n"
        f"📦 <b>Вантаж:</b> {html_escape(offer['cargo'])}\n"
        f"⚖️ <b>Вага:</b> {fmt_weight(offer['weight_t'])} т"
        f"{date}{extra}{close_at}"
    )


def format_offer_for_carrier(
    offer: dict, request_id: int, proposal: dict
) -> str:
    status = status_emoji(offer["status"]) + " " + status_text(offer["status"])
    extra = (
        f"\n💬 {html_escape(offer['extra_info'])}"
        if offer.get("extra_info")
        else ""
    )
    date = (
        f"\n📅 Завантаження: {html_escape(offer['load_date'])}"
        if offer.get("load_date")
        else ""
    )
    price_vat = (
        f"\n• З ПДВ: <b>{fmt_price(proposal['price_with_vat'])} грн</b>"
        if proposal.get("price_with_vat")
        else ""
    )
    price_no_vat = (
        f"\n• Без ПДВ: <b>{fmt_price(proposal['price_without_vat'])} грн</b>"
        if proposal.get("price_without_vat")
        else ""
    )
    your_prices = ""
    if price_vat or price_no_vat:
        your_prices = f"\n\n<b>Ваша пропозиція:</b>{price_vat}{price_no_vat}"
    return (
        f"<b>Заявка #{request_id}</b> — Оголошення #{offer['id']} — {status}\n"
        f"━━━━━━━━━━━━━━\n"
        f"🚛 {html_escape(offer['route_from'])} → "
        f"{html_escape(offer['route_to'])}\n"
        f"📦 {html_escape(offer['cargo'])}\n"
        f"⚖️ {fmt_weight(offer['weight_t'])} т"
        f"{date}{extra}"
        f"{your_prices}"
    )


# ────────────────────  Клавіатури  ────────────────────
def admin_menu_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    pending_count = len(db.list_pending_users())
    pending_label = ADMIN_BTN_PENDING + (
        f" ({pending_count})" if pending_count else ""
    )
    rows = [
        [
            KeyboardButton(text=ADMIN_BTN_NEW),
            KeyboardButton(text=ADMIN_BTN_LIST),
        ],
        [
            KeyboardButton(text=pending_label),
            KeyboardButton(text=ADMIN_BTN_STATS),
        ],
    ]
    if is_super_admin(user_id):
        rows.append(
            [
                KeyboardButton(text=ADMIN_BTN_ADMINS),
                KeyboardButton(text=ADMIN_BTN_BROADCAST),
            ]
        )
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def carrier_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=CARRIER_BTN_MY_PROFILE)],
            [KeyboardButton(text=CARRIER_BTN_HELP)],
        ],
        resize_keyboard=True,
    )


def carrier_card_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text=BTN_WITH_VAT),
                KeyboardButton(text=BTN_WITHOUT_VAT),
            ],
            [KeyboardButton(text=BTN_CONTACT)],
            [KeyboardButton(text=BTN_CANCEL)],
        ],
        resize_keyboard=True,
    )


def price_input_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=BTN_CANCEL)]],
        resize_keyboard=True,
    )


def new_offer_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=BTN_CANCEL)]],
        resize_keyboard=True,
    )


def new_offer_skip_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=BTN_SKIP)],
            [KeyboardButton(text=BTN_CANCEL)],
        ],
        resize_keyboard=True,
    )


def confirm_offer_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=BTN_CONFIRM)],
            [KeyboardButton(text=BTN_CANCEL)],
        ],
        resize_keyboard=True,
    )


def share_phone_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=BTN_SHARE_PHONE, request_contact=True)],
            [KeyboardButton(text=BTN_CANCEL)],
        ],
        resize_keyboard=True,
    )


def channel_offer_keyboard(offer_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[[
            InlineKeyboardButton(
                text="💵 Ввести вашу пропозицію",
                url=f"https://t.me/{BOT_USERNAME}?start=offer_{offer_id}",
            )
        ]]
    )


def offer_actions_inline(offer: dict) -> InlineKeyboardMarkup:
    buttons = [
        [
            InlineKeyboardButton(
                text="📊 Звіт", callback_data=f"report:{offer['id']}"
            ),
            InlineKeyboardButton(
                text="📥 Excel", callback_data=f"excel:{offer['id']}"
            ),
        ],
        [
            InlineKeyboardButton(
                text="✏️ Редагувати", callback_data=f"edit:{offer['id']}"
            ),
        ],
    ]
    if offer["status"] == "closed":
        buttons.append([
            InlineKeyboardButton(
                text="🟢 Відкрити знов", callback_data=f"reopen:{offer['id']}"
            )
        ])
    else:
        buttons.append([
            InlineKeyboardButton(
                text="🔴 Закрити", callback_data=f"close:{offer['id']}"
            )
        ])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def filter_keyboard(active: str) -> InlineKeyboardMarkup:
    """Рядок-фільтр над списком оголошень."""

    def lbl(name, value):
        return ("🔹 " if value == active else "") + name

    return InlineKeyboardMarkup(
        inline_keyboard=[[
            InlineKeyboardButton(
                text=lbl("Усі", "all"), callback_data="flt:all"
            ),
            InlineKeyboardButton(
                text=lbl("🟢", "open"), callback_data="flt:open"
            ),
            InlineKeyboardButton(
                text=lbl("🟡", "in_progress"), callback_data="flt:in_progress"
            ),
            InlineKeyboardButton(
                text=lbl("🔴", "closed"), callback_data="flt:closed"
            ),
        ]]
    )


def edit_field_keyboard(offer_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="🚛 Маршрут ВІД",
                    callback_data=f"editf:{offer_id}:route_from",
                ),
                InlineKeyboardButton(
                    text="🚛 Маршрут ДО",
                    callback_data=f"editf:{offer_id}:route_to",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="📦 Вантаж",
                    callback_data=f"editf:{offer_id}:cargo",
                ),
                InlineKeyboardButton(
                    text="⚖️ Вага",
                    callback_data=f"editf:{offer_id}:weight_t",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="📅 Дата",
                    callback_data=f"editf:{offer_id}:load_date",
                ),
                InlineKeyboardButton(
                    text="💬 Примітка",
                    callback_data=f"editf:{offer_id}:extra_info",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="👤 Контакт (ім'я)",
                    callback_data=f"editf:{offer_id}:contact_name",
                ),
                InlineKeyboardButton(
                    text="📞 Контакт (тел)",
                    callback_data=f"editf:{offer_id}:contact_phone",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="❌ Закрити меню",
                    callback_data=f"editcancel:{offer_id}",
                ),
            ],
        ]
    )


def admins_list_inline(
    admins: list, viewer_is_super: bool
) -> InlineKeyboardMarkup:
    rows = []
    for a in admins:
        label = f"🗑 {a['first_name'] or a['username'] or a['user_id']}"
        rows.append([
            InlineKeyboardButton(
                text=label, callback_data=f"rmadmin:{a['user_id']}"
            )
        ])
    if viewer_is_super:
        rows.append([
            InlineKeyboardButton(
                text="➕ Додати адміна", callback_data="addadmin"
            )
        ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def proposal_action_inline(
    offer_id: int, proposal: dict, is_winner: bool
) -> InlineKeyboardMarkup:
    rows = []
    if not is_winner:
        rows.append([
            InlineKeyboardButton(
                text="✅ Обрати переможцем",
                callback_data=f"pickwin:{offer_id}:{proposal['id']}",
            )
        ])
    u = db.get_user(proposal["user_id"])
    if u and u.get("is_blacklisted"):
        rows.append([
            InlineKeyboardButton(
                text="↩️ Розблокувати",
                callback_data=f"unban:{proposal['user_id']}:{offer_id}",
            )
        ])
    else:
        rows.append([
            InlineKeyboardButton(
                text="🚫 В чорний список",
                callback_data=f"ban:{proposal['user_id']}:{offer_id}",
            )
        ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def broadcast_confirm_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📤 Розіслати")],
            [KeyboardButton(text=BTN_CANCEL)],
        ],
        resize_keyboard=True,
    )


# ────────────────────  Публікація в канал  ────────────────────
async def publish_offer_to_channel(offer: dict) -> int:
    """Публікує оголошення в канал. Повертає message_id."""
    text = format_offer_for_channel(offer)
    kb = channel_offer_keyboard(offer["id"])
    if offer.get("photo_file_id"):
        msg = await bot.send_photo(
            CHANNEL_ID,
            photo=offer["photo_file_id"],
            caption=text,
            reply_markup=kb,
        )
    else:
        msg = await bot.send_message(CHANNEL_ID, text, reply_markup=kb)
    return msg.message_id


async def update_channel_post(offer: dict):
    """Оновлює пост у каналі (текст і статус) із reply-кнопкою."""
    if not offer.get("channel_message_id"):
        return
    text = format_offer_for_channel(offer)
    kb = channel_offer_keyboard(offer["id"]) if offer["status"] != "closed" else None
    try:
        if offer.get("photo_file_id"):
            await bot.edit_message_caption(
                chat_id=CHANNEL_ID,
                message_id=offer["channel_message_id"],
                caption=text,
                reply_markup=kb,
            )
        else:
            await bot.edit_message_text(
                text,
                chat_id=CHANNEL_ID,
                message_id=offer["channel_message_id"],
                reply_markup=kb,
            )
    except TelegramBadRequest as e:
        if "message is not modified" not in str(e):
            logger.warning("Не вдалось оновити пост: %s", e)


# ────────────────────  Сповіщення адмінам  ────────────────────
async def notify_admins(text: str, **kwargs):
    """Пересилає повідомлення всім адмінам (супер + з БД)."""
    ids = {SUPER_ADMIN_ID}
    for a in db.list_admins():
        ids.add(a["user_id"])
    for uid in ids:
        try:
            await bot.send_message(uid, text, **kwargs)
        except (TelegramForbiddenError, TelegramBadRequest) as e:
            logger.warning("Не вдалось повідомити адміна %s: %s", uid, e)


async def notify_admins_new_price(
    offer: dict, proposal: dict, user_id: int
):
    u = db.get_user(user_id) or {}
    link = user_profile_link(
        user_id, proposal.get("username"), proposal.get("first_name")
    )
    wins = u.get("wins_count", 0)
    total = u.get("total_proposals", 0)
    phone = u.get("phone") or proposal.get("phone") or "—"
    edrpou = u.get("edrpou") or "—"
    full = html_escape(u.get("full_name") or "")
    price_vat = (
        f"💰 З ПДВ: <b>{fmt_price(proposal['price_with_vat'])} грн</b>\n"
        if proposal.get("price_with_vat")
        else ""
    )
    price_no_vat = (
        f"💰 Без ПДВ: <b>{fmt_price(proposal['price_without_vat'])} грн</b>\n"
        if proposal.get("price_without_vat")
        else ""
    )
    text = (
        f"🔔 <b>Нова ціна</b> на оголошення #{offer['id']}\n"
        f"Маршрут: {html_escape(offer['route_from'])} → "
        f"{html_escape(offer['route_to'])}\n\n"
        f"{price_vat}{price_no_vat}\n"
        f"Перевізник: {link}"
        + (f" ({full})" if full else "")
        + f"\n📞 {html_escape(phone)}"
        f"\n🧾 ЄДРПОУ/ІПН: {html_escape(edrpou)}"
        f"\n📈 Історія: {wins} перемог / {total} заявок"
    )
    await notify_admins(text)


# ────────────────────  /start  ────────────────────
@router.message(CommandStart(deep_link=True))
async def cmd_start_deeplink(
    message: Message, command: CommandObject, state: FSMContext
):
    db.upsert_user(
        message.from_user.id,
        message.from_user.username or "",
        message.from_user.first_name or "",
    )
    await state.clear()

    arg = command.args or ""
    if not arg.startswith("offer_"):
        await cmd_start_regular(message, state)
        return

    try:
        offer_id = int(arg.split("_", 1)[1])
    except (ValueError, IndexError):
        await cmd_start_regular(message, state)
        return

    offer = db.get_offer(offer_id)
    if not offer:
        await message.answer("Вибачте, оголошення не знайдено.")
        return

    # Адмін? — покажемо адмін-меню й просто покажемо звіт-посилання.
    if is_admin(message.from_user.id):
        await message.answer(
            f"Ви адміністратор. Ось ваше меню. "
            f"Оголошення #{offer_id} можна переглянути у '{ADMIN_BTN_LIST}'.",
            reply_markup=admin_menu_keyboard(message.from_user.id),
        )
        return

    # Blacklist?
    if db.is_blacklisted(message.from_user.id):
        u = db.get_user(message.from_user.id) or {}
        reason = u.get("blacklist_reason") or ""
        await message.answer(
            "🚫 На жаль, ви заблоковані і не можете подавати пропозиції."
            + (f"\nПричина: {html_escape(reason)}" if reason else ""),
            reply_markup=ReplyKeyboardRemove(),
        )
        return

    # Pending?
    if db.is_pending(message.from_user.id):
        await message.answer(
            "⏳ Ваша заявка на реєстрацію ще на перевірці у адміністратора. "
            "Як тільки її схвалять — ви отримаєте сповіщення й зможете "
            "повернутись до цього оголошення.",
            reply_markup=ReplyKeyboardRemove(),
        )
        return

    # Не зареєстрований → стартуємо реєстрацію
    if not db.is_registered(message.from_user.id):
        await state.update_data(pending_offer_id=offer_id)
        await start_registration(message, state)
        return

    # Реєстрований — створюємо/беремо заявку
    request_id, proposal = db.get_or_create_proposal(
        offer_id,
        message.from_user.id,
        message.from_user.username or "",
        message.from_user.first_name or "",
    )
    await state.update_data(
        active_offer_id=offer_id, active_request_id=request_id
    )

    if offer.get("photo_file_id"):
        try:
            await bot.send_photo(
                message.chat.id, photo=offer["photo_file_id"]
            )
        except TelegramBadRequest:
            pass

    await message.answer(
        format_offer_for_carrier(offer, request_id, proposal),
        reply_markup=carrier_card_keyboard(),
    )


@router.message(CommandStart())
async def cmd_start_regular(message: Message, state: FSMContext):
    db.upsert_user(
        message.from_user.id,
        message.from_user.username or "",
        message.from_user.first_name or "",
    )
    await state.clear()

    if is_admin(message.from_user.id):
        await message.answer(
            "👋 Адмін-панель. Оберіть дію нижче.",
            reply_markup=admin_menu_keyboard(message.from_user.id),
        )
        return

    if db.is_blacklisted(message.from_user.id):
        u = db.get_user(message.from_user.id) or {}
        reason = u.get("blacklist_reason") or ""
        await message.answer(
            "🚫 Ви заблоковані і не можете подавати пропозиції."
            + (f"\nПричина: {html_escape(reason)}" if reason else ""),
            reply_markup=ReplyKeyboardRemove(),
        )
        return

    if db.is_pending(message.from_user.id):
        await message.answer(
            "⏳ Ваша заявка на реєстрацію ще на перевірці у адміністратора. "
            "Дочекайтесь сповіщення про результат.",
            reply_markup=ReplyKeyboardRemove(),
        )
        return

    if not db.is_registered(message.from_user.id):
        await start_registration(message, state)
        return

    u = db.get_user(message.from_user.id) or {}
    await message.answer(
        f"👋 Вітаємо, {html_escape(u.get('full_name') or '')}!\n"
        "Ви побачите нові оголошення у каналі. Натискайте кнопку під "
        "постом — і бот попросить ввести вашу ціну.",
        reply_markup=carrier_menu_keyboard(),
    )


# ────────────────────  Реєстрація перевізника  ────────────────────
async def start_registration(message: Message, state: FSMContext):
    await state.set_state(RegisterStates.full_name)
    await message.answer(
        "👋 Щоб подавати пропозиції на перевезення, треба один раз "
        "зареєструватись. Ваша заявка буде перевірена адміністратором.\n\n"
        "<b>Крок 1/3.</b> Введіть ваше <b>ПІБ</b> (Прізвище Ім'я По-батькові) "
        "або назву ФОП/ТОВ:",
        reply_markup=new_offer_keyboard(),
    )


@router.message(RegisterStates.full_name)
async def reg_full_name(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        await _universal_cancel(message, state)
        return
    text = (message.text or "").strip()
    if len(text) < 3 or len(text) > 120:
        await message.answer(
            "Введіть ПІБ довжиною 3-120 символів. Наприклад: "
            "<code>Петренко Іван Миколайович</code> або <code>ТОВ Рух-Транс</code>"
        )
        return
    await state.update_data(full_name=text)
    await state.set_state(RegisterStates.phone)
    await message.answer(
        "<b>Крок 2/3.</b> Надішліть ваш <b>телефон</b>. Можна натиснути "
        "кнопку «📞 Поділитись номером» або ввести вручну у форматі "
        "<code>+380501234567</code>.",
        reply_markup=share_phone_keyboard(),
    )


@router.message(RegisterStates.phone, F.contact)
async def reg_phone_contact(message: Message, state: FSMContext):
    phone = parse_phone(message.contact.phone_number)
    if not phone:
        await message.answer("Не вдалось обробити номер. Введіть вручну.")
        return
    await _reg_phone_save(message, state, phone)


@router.message(RegisterStates.phone)
async def reg_phone_text(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        await _universal_cancel(message, state)
        return
    phone = parse_phone(message.text or "")
    if not phone:
        await message.answer(
            "Некоректний номер. Приклад: <code>+380501234567</code> або "
            "<code>0501234567</code>. Спробуйте ще раз."
        )
        return
    await _reg_phone_save(message, state, phone)


async def _reg_phone_save(
    message: Message, state: FSMContext, phone: str
):
    await state.update_data(phone=phone)
    await state.set_state(RegisterStates.edrpou)
    await message.answer(
        f"✅ Номер збережено: <code>{phone}</code>\n\n"
        "<b>Крок 3/3.</b> Введіть <b>ЄДРПОУ</b> (для ТОВ, 8 цифр) або "
        "<b>ІПН/РНОКПП</b> (для ФОП, 10 цифр). Можна пропустити, якщо не маєте.",
        reply_markup=new_offer_skip_keyboard(),
    )


def _pending_review_keyboard(user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[[
            InlineKeyboardButton(
                text="✅ Схвалити", callback_data=f"approve:{user_id}"
            ),
            InlineKeyboardButton(
                text="❌ Відхилити", callback_data=f"reject:{user_id}"
            ),
        ]]
    )


async def _notify_admins_new_registration(user_id: int):
    u = db.get_user(user_id) or {}
    link = user_profile_link(
        user_id, u.get("username"), u.get("first_name")
    )
    text = (
        "📝 <b>Нова заявка на реєстрацію</b>\n\n"
        f"Користувач: {link}\n"
        f"ПІБ / Назва: <b>{html_escape(u.get('full_name'))}</b>\n"
        f"Телефон: <code>{html_escape(u.get('phone'))}</code>\n"
        f"ЄДРПОУ / ІПН: <code>{html_escape(u.get('edrpou') or '—')}</code>"
    )
    ids = {SUPER_ADMIN_ID}
    for a in db.list_admins():
        ids.add(a["user_id"])
    for uid in ids:
        try:
            await bot.send_message(
                uid, text, reply_markup=_pending_review_keyboard(user_id)
            )
        except (TelegramForbiddenError, TelegramBadRequest) as e:
            logger.warning(
                "Не вдалось повідомити адміна %s про реєстрацію: %s", uid, e
            )


@router.message(RegisterStates.edrpou)
async def reg_edrpou(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        await _universal_cancel(message, state)
        return
    if message.text == BTN_SKIP:
        edrpou = None
    else:
        edrpou = parse_edrpou(message.text or "")
        if not edrpou:
            await message.answer(
                "ЄДРПОУ має бути 8 цифр, а ІПН — 10. Введіть ще раз або натисніть «Пропустити»."
            )
            return

    data = await state.get_data()
    db.submit_registration(
        message.from_user.id,
        full_name=data["full_name"],
        phone=data["phone"],
        edrpou=edrpou,
    )

    await _notify_admins_new_registration(message.from_user.id)

    await state.clear()
    await message.answer(
        "✅ <b>Заявку на реєстрацію подано.</b>\n\n"
        "Адміністратор перевірить ваші дані і ви отримаєте повідомлення "
        "про результат. Зазвичай це займає до кількох годин.\n\n"
        "Дякуємо за терпіння!",
        reply_markup=ReplyKeyboardRemove(),
    )


# ────────────────────  Універсальний «Скасувати»  ────────────────────
async def _universal_cancel(message: Message, state: FSMContext):
    current = await state.get_state()
    await state.clear()
    if is_admin(message.from_user.id):
        await message.answer(
            "Скасовано.", reply_markup=admin_menu_keyboard(message.from_user.id)
        )
    else:
        # Якщо перевізник ще не зареєстрований — повертаємо на start
        if not db.is_registered(message.from_user.id):
            await message.answer(
                "Скасовано. Напишіть /start коли будете готові зареєструватись.",
                reply_markup=ReplyKeyboardRemove(),
            )
        else:
            await message.answer(
                "Скасовано.", reply_markup=carrier_menu_keyboard()
            )


@router.message(Command("cancel"))
async def cmd_cancel(message: Message, state: FSMContext):
    await _universal_cancel(message, state)


@router.message(Command("reset"))
async def cmd_reset(message: Message, state: FSMContext):
    await _universal_cancel(message, state)


@router.message(F.text == BTN_CANCEL)
async def btn_cancel(message: Message, state: FSMContext):
    await _universal_cancel(message, state)


# ────────────────────  Профіль перевізника  ────────────────────
@router.message(F.text == CARRIER_BTN_MY_PROFILE)
async def btn_my_profile(message: Message):
    u = db.get_user(message.from_user.id)
    if not u or not u.get("is_registered"):
        if u and u.get("is_pending"):
            await message.answer(
                "⏳ Ваша заявка на реєстрацію ще на перевірці."
            )
            return
        await message.answer("Ви ще не зареєстровані. Напишіть /start.")
        return
    text = (
        "<b>👤 Ваш профіль</b>\n"
        f"ПІБ / Назва: <b>{html_escape(u.get('full_name'))}</b>\n"
        f"Телефон: <code>{html_escape(u.get('phone'))}</code>\n"
        f"ЄДРПОУ / ІПН: <code>{html_escape(u.get('edrpou') or '—')}</code>\n"
        f"Історія: <b>{u.get('wins_count', 0)}</b> перемог з "
        f"<b>{u.get('total_proposals', 0)}</b> заявок"
    )
    await message.answer(text)


@router.message(F.text == CARRIER_BTN_HELP)
async def btn_carrier_help(message: Message):
    await message.answer(
        "<b>❓ Як користуватись</b>\n\n"
        "1. Ви побачите нові оголошення в каналі.\n"
        "2. Під кожним постом — кнопка «💵 Ввести вашу пропозицію».\n"
        "3. Натисніть, відкриється цей бот з карткою заявки.\n"
        "4. Введіть тариф з ПДВ і/або без ПДВ.\n"
        "5. Адміністратор побачить вашу ціну миттєво.\n\n"
        "Ціну можна оновлювати — останнє значення йде у фінальний звіт.\n"
        "Кнопка «📱 Передати контакт» — надіслати адміну ваш телефон."
    )


# ────────────────────  Створення оголошення  ────────────────────
async def _start_new_offer(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(NewOfferStates.route_from)
    await message.answer(
        "<b>Нове оголошення — крок 1/9.</b>\n"
        "Маршрут <b>ВІД</b> (звідки вантаж, напр. <code>Київ, Київська обл.</code>):",
        reply_markup=new_offer_keyboard(),
    )


@router.message(Command("new"))
async def cmd_new(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await _start_new_offer(message, state)


@router.message(F.text == ADMIN_BTN_NEW)
async def btn_new(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await _start_new_offer(message, state)


@router.message(NewOfferStates.route_from)
async def new_route_from(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    await state.update_data(route_from=message.text.strip())
    await state.set_state(NewOfferStates.route_to)
    await message.answer("Крок 2/9. Маршрут <b>ДО</b>:")


@router.message(NewOfferStates.route_to)
async def new_route_to(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    await state.update_data(route_to=message.text.strip())
    await state.set_state(NewOfferStates.cargo)
    await message.answer(
        "Крок 3/9. <b>Вантаж</b> (напр. <code>Пшениця 3-го класу</code>):"
    )


@router.message(NewOfferStates.cargo)
async def new_cargo(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    await state.update_data(cargo=message.text.strip())
    await state.set_state(NewOfferStates.weight)
    await message.answer("Крок 4/9. <b>Вага, тонн</b> (напр. <code>25</code>):")


@router.message(NewOfferStates.weight)
async def new_weight(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    w = parse_tonnage(message.text or "")
    if not w:
        await message.answer("Введіть число тонн від 0 до 100000 (напр. 99777).")
        return
    await state.update_data(weight_t=w)
    await state.set_state(NewOfferStates.load_date)
    await message.answer(
        "Крок 5/9. <b>Дата завантаження</b> (текстом, напр. "
        "<code>25.04.2026</code>). Можна пропустити.",
        reply_markup=new_offer_skip_keyboard(),
    )


@router.message(NewOfferStates.load_date)
async def new_load_date(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    date = "" if message.text == BTN_SKIP else message.text.strip()
    await state.update_data(load_date=date)
    await state.set_state(NewOfferStates.extra_info)
    await message.answer(
        "Крок 6/9. <b>Додаткова інформація</b> (особливості, вимоги до авто тощо). "
        "Можна пропустити.",
        reply_markup=new_offer_skip_keyboard(),
    )


@router.message(NewOfferStates.extra_info)
async def new_extra(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    extra = "" if message.text == BTN_SKIP else message.text.strip()
    await state.update_data(extra_info=extra)
    await state.set_state(NewOfferStates.contact_name)
    await message.answer(
        "Крок 7/9. <b>Контактна особа</b> (кому передзвонить перевізник, якщо захочете). "
        "Можна пропустити.",
        reply_markup=new_offer_skip_keyboard(),
    )


@router.message(NewOfferStates.contact_name)
async def new_contact_name(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    cname = "" if message.text == BTN_SKIP else message.text.strip()
    await state.update_data(contact_name=cname)
    await state.set_state(NewOfferStates.contact_phone)
    await message.answer(
        "Крок 8/9. <b>Телефон контактної особи</b>. Можна пропустити.",
        reply_markup=new_offer_skip_keyboard(),
    )


@router.message(NewOfferStates.contact_phone)
async def new_contact_phone(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    if message.text == BTN_SKIP:
        await state.update_data(contact_phone="")
    else:
        phone = parse_phone(message.text or "")
        if not phone:
            await message.answer(
                "Некоректний телефон. Наприклад: <code>+380501234567</code>. "
                "Або натисніть «Пропустити»."
            )
            return
        await state.update_data(contact_phone=phone)
    await state.set_state(NewOfferStates.photo)
    await message.answer(
        "Крок 9/9. <b>Фото вантажу</b> (надішліть зображення). Можна пропустити.",
        reply_markup=new_offer_skip_keyboard(),
    )


@router.message(NewOfferStates.photo, F.photo)
async def new_photo(message: Message, state: FSMContext):
    file_id = message.photo[-1].file_id
    await state.update_data(photo_file_id=file_id)
    await _new_offer_confirm(message, state)


@router.message(NewOfferStates.photo)
async def new_photo_text(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    if message.text == BTN_SKIP:
        await state.update_data(photo_file_id=None)
        await _new_offer_confirm(message, state)
        return
    await message.answer(
        "Надішліть фото або натисніть «Пропустити»."
    )


async def _new_offer_confirm(message: Message, state: FSMContext):
    data = await state.get_data()
    preview = {
        "id": "—",
        "status": "open",
        "route_from": data.get("route_from"),
        "route_to": data.get("route_to"),
        "cargo": data.get("cargo"),
        "weight_t": data.get("weight_t"),
        "load_date": data.get("load_date"),
        "extra_info": data.get("extra_info"),
    }
    contact_block = ""
    if data.get("contact_name") or data.get("contact_phone"):
        contact_block = (
            f"\n\n👤 Контакт: {html_escape(data.get('contact_name') or '—')}"
            f" · {html_escape(data.get('contact_phone') or '—')}"
        )
    auto_close = ""
    if AUTO_CLOSE_HOURS > 0:
        close_at = (
            datetime.utcnow() + timedelta(hours=AUTO_CLOSE_HOURS)
        ).strftime("%Y-%m-%d %H:%M:%S")
        auto_close = f"\n\n⏳ Автозакриття через {AUTO_CLOSE_HOURS} год"
    await state.set_state(NewOfferStates.confirm)
    await message.answer(
        "<b>Попередній перегляд:</b>\n\n"
        + format_offer_for_channel(preview)
        + contact_block
        + auto_close
        + "\n\nОпублікувати?",
        reply_markup=confirm_offer_keyboard(),
    )


@router.message(NewOfferStates.confirm)
async def new_confirm(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    if message.text != BTN_CONFIRM:
        await message.answer("Натисніть «✅ Опублікувати» або «❌ Скасувати».")
        return

    data = await state.get_data()
    auto_close_at = None
    if AUTO_CLOSE_HOURS > 0:
        auto_close_at = (
            datetime.utcnow() + timedelta(hours=AUTO_CLOSE_HOURS)
        ).strftime("%Y-%m-%d %H:%M:%S")

    offer_id = db.create_offer(
        route_from=data["route_from"],
        route_to=data["route_to"],
        cargo=data["cargo"],
        weight_t=data["weight_t"],
        load_date=data.get("load_date", ""),
        extra_info=data.get("extra_info", ""),
        contact_name=data.get("contact_name", ""),
        contact_phone=data.get("contact_phone", ""),
        photo_file_id=data.get("photo_file_id"),
        auto_close_at=auto_close_at,
        created_by=message.from_user.id,
    )
    offer = db.get_offer(offer_id)

    try:
        msg_id = await publish_offer_to_channel(offer)
        db.set_offer_message_id(offer_id, msg_id)
    except Exception as e:
        logger.exception("Не вдалось опублікувати в канал")
        await message.answer(
            f"⚠️ Оголошення #{offer_id} збережено у БД, але не опубліковано в "
            f"каналі: <code>{html_escape(str(e))}</code>. Перевірте права бота "
            "та CHANNEL_ID.",
            reply_markup=admin_menu_keyboard(message.from_user.id),
        )
        await state.clear()
        return

    await state.clear()
    offer = db.get_offer(offer_id)
    await message.answer(
        f"✅ <b>Оголошення #{offer_id} опубліковане.</b>\n"
        f"<a href=\"{channel_post_url(msg_id)}\">Відкрити у каналі</a>",
        reply_markup=admin_menu_keyboard(message.from_user.id),
    )


# ────────────────────  Список оголошень + фільтри  ────────────────────
async def _send_list(
    message: Message, status_filter: str = "active", edit: bool = False
):
    if status_filter == "all":
        offers = db.list_offers()
    elif status_filter == "active":
        offers = db.list_offers(["open", "in_progress"])
    else:
        offers = db.list_offers([status_filter])

    if not offers:
        text = "Оголошень немає."
    else:
        lines = [f"<b>Оголошення ({len(offers)}):</b>\n"]
        for o in offers[:30]:
            cnt = db.count_proposals(o["id"])
            best_vat = "—"
            if cnt:
                props = db.list_proposals(o["id"])
                with_vat_prices = [
                    p["price_with_vat"] for p in props if p["price_with_vat"]
                ]
                if with_vat_prices:
                    best_vat = f"{fmt_price(min(with_vat_prices))} грн"
            lines.append(
                f"{status_emoji(o['status'])} #{o['id']} "
                f"{html_escape(o['route_from'])} → {html_escape(o['route_to'])} · "
                f"{fmt_weight(o['weight_t'])}т · "
                f"пропозицій: <b>{cnt}</b> · найкраща з ПДВ: {best_vat}"
            )
        text = "\n".join(lines)

    kb = filter_keyboard(status_filter if status_filter != "active" else "all")
    if edit:
        try:
            await message.edit_text(text, reply_markup=kb)
        except TelegramBadRequest:
            await message.answer(text, reply_markup=kb)
    else:
        await message.answer(text, reply_markup=kb)

    # Після списку — окремі повідомлення з inline-кнопками для кожного
    if not edit and offers:
        for o in offers[:10]:
            await message.answer(
                f"#{o['id']} · {status_emoji(o['status'])} {status_text(o['status'])}",
                reply_markup=offer_actions_inline(o),
            )


@router.message(Command("list"))
async def cmd_list(message: Message):
    if not is_admin(message.from_user.id):
        return
    await _send_list(message)


@router.message(F.text == ADMIN_BTN_LIST)
async def btn_list(message: Message):
    if not is_admin(message.from_user.id):
        return
    await _send_list(message)


@router.callback_query(F.data.startswith("flt:"))
async def cb_filter(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    status = cb.data.split(":", 1)[1]
    await _send_list(cb.message, status, edit=True)
    await cb.answer()


# ────────────────────  Звіт  ────────────────────
async def _send_report(target, offer_id: int):
    offer = db.get_offer(offer_id)
    if not offer:
        await target.answer("Оголошення не знайдено.")
        return
    proposals = db.list_proposals(offer_id)

    header = (
        f"<b>Звіт по оголошенню #{offer_id}</b>\n"
        f"{status_emoji(offer['status'])} {status_text(offer['status'])} · "
        f"{html_escape(offer['route_from'])} → {html_escape(offer['route_to'])}\n"
        f"📦 {html_escape(offer['cargo'])} · ⚖️ {fmt_weight(offer['weight_t'])} т"
    )
    if offer.get("channel_message_id"):
        header += (
            f"\n<a href=\"{channel_post_url(offer['channel_message_id'])}\">"
            "Відкрити у каналі</a>"
        )
    await target.answer(header)

    if not proposals:
        await target.answer("Ще немає пропозицій.")
        return

    # Підсумок: найкращі ціни
    with_vat = [p for p in proposals if p.get("price_with_vat")]
    without_vat = [p for p in proposals if p.get("price_without_vat")]

    summary = "<b>Найкращі ціни:</b>"
    if with_vat:
        best = min(with_vat, key=lambda p: p["price_with_vat"])
        link = user_profile_link(
            best["user_id"], best.get("username"), best.get("first_name")
        )
        summary += (
            f"\n💰 З ПДВ: <b>{fmt_price(best['price_with_vat'])} грн</b> — {link}"
        )
    if without_vat:
        best = min(without_vat, key=lambda p: p["price_without_vat"])
        link = user_profile_link(
            best["user_id"], best.get("username"), best.get("first_name")
        )
        summary += (
            f"\n💰 Без ПДВ: <b>{fmt_price(best['price_without_vat'])} грн</b> — {link}"
        )
    if not with_vat and not without_vat:
        summary += "\n—"
    await target.answer(summary)

    # Кожна пропозиція окремим повідомленням з кнопками
    winner_proposal_id = offer.get("winner_proposal_id")
    for i, p in enumerate(proposals, 1):
        u = db.get_user(p["user_id"]) or {}
        link = user_profile_link(
            p["user_id"], p.get("username"), p.get("first_name")
        )
        phone = u.get("phone") or p.get("phone") or "—"
        full = html_escape(u.get("full_name") or "")
        edrpou = u.get("edrpou") or "—"
        bl = " 🚫" if u.get("is_blacklisted") else ""
        win = " 🏆 Переможець" if p["id"] == winner_proposal_id else ""
        text = (
            f"<b>Заявка #{p['id']}</b>{win}{bl}\n"
            f"Перевізник: {link}"
            + (f" ({full})" if full else "")
            + f"\n📞 <code>{html_escape(phone)}</code>"
            f"\n🧾 ЄДРПОУ/ІПН: {html_escape(edrpou)}"
            f"\n📈 {u.get('wins_count', 0)} перемог / "
            f"{u.get('total_proposals', 0)} заявок\n"
            + (
                f"💰 З ПДВ: <b>{fmt_price(p['price_with_vat'])} грн</b>\n"
                if p.get("price_with_vat")
                else ""
            )
            + (
                f"💰 Без ПДВ: <b>{fmt_price(p['price_without_vat'])} грн</b>"
                if p.get("price_without_vat")
                else ""
            )
        )
        kb = None
        if offer["status"] != "closed":
            kb = proposal_action_inline(
                offer_id, p, p["id"] == winner_proposal_id
            )
        await target.answer(text, reply_markup=kb)


@router.message(Command("report"))
async def cmd_report(message: Message, command: CommandObject):
    if not is_admin(message.from_user.id):
        return
    if not command.args:
        await message.answer("Напр. <code>/report 5</code>")
        return
    try:
        oid = int(command.args.strip())
    except ValueError:
        await message.answer("ID — число.")
        return
    await _send_report(message, oid)


@router.callback_query(F.data.startswith("report:"))
async def cb_report(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    oid = int(cb.data.split(":", 1)[1])
    await _send_report(cb.message, oid)
    await cb.answer()


# ────────────────────  Експорт в Excel  ────────────────────
@router.callback_query(F.data.startswith("excel:"))
async def cb_excel(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    oid = int(cb.data.split(":", 1)[1])
    offer = db.get_offer(oid)
    if not offer:
        await cb.answer("Не знайдено")
        return
    await cb.answer("Готую Excel…")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Оголошення {oid}"

    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DDE6F0")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Шапка
    ws["A1"] = f"Оголошення #{oid}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    meta = [
        ("Статус", f"{status_emoji(offer['status'])} {status_text(offer['status'])}"),
        ("Маршрут", f"{offer['route_from']} → {offer['route_to']}"),
        ("Вантаж", offer["cargo"]),
        ("Вага, т", fmt_weight(offer["weight_t"])),
        ("Дата завантаження", offer.get("load_date") or "—"),
        ("Примітка", offer.get("extra_info") or "—"),
        ("Контактна особа", offer.get("contact_name") or "—"),
        ("Телефон контакту", offer.get("contact_phone") or "—"),
        ("Створено", offer.get("created_at") or "—"),
    ]
    row = 3
    for k, v in meta:
        ws.cell(row=row, column=1, value=k).font = bold
        ws.cell(row=row, column=2, value=str(v))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 1

    # Таблиця пропозицій
    row += 1
    headers = [
        "№ заявки", "ПІБ / Назва", "Username", "Телефон",
        "ЄДРПОУ/ІПН", "З ПДВ", "Без ПДВ",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = bold
        cell.fill = hdr_fill
    row += 1

    proposals = db.list_proposals(oid)
    winner_id = offer.get("winner_proposal_id")
    for p in proposals:
        u = db.get_user(p["user_id"]) or {}
        mark = " 🏆" if p["id"] == winner_id else ""
        ws.cell(row=row, column=1, value=f"#{p['id']}{mark}")
        ws.cell(row=row, column=2, value=u.get("full_name") or p.get("first_name") or "")
        ws.cell(row=row, column=3, value=f"@{p['username']}" if p.get("username") else "")
        ws.cell(row=row, column=4, value=u.get("phone") or p.get("phone") or "")
        ws.cell(row=row, column=5, value=u.get("edrpou") or "")
        ws.cell(row=row, column=6, value=p.get("price_with_vat") or "")
        ws.cell(row=row, column=7, value=p.get("price_without_vat") or "")
        row += 1

    # Ширини
    widths = [10, 26, 16, 18, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in ws.iter_rows(min_row=1, max_row=row - 1):
        for c in r:
            c.alignment = wrap

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"offer_{oid}.xlsx"
    await cb.message.answer_document(
        BufferedInputFile(buf.read(), filename=filename),
        caption=f"Звіт по оголошенню #{oid}",
    )


# ────────────────────  Вибір переможця / blacklist  ────────────────────
@router.callback_query(F.data.startswith("pickwin:"))
async def cb_pick_winner(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    _, oid_s, pid_s = cb.data.split(":")
    oid, pid = int(oid_s), int(pid_s)
    offer = db.get_offer(oid)
    proposal = db.get_proposal(pid)
    if not offer or not proposal:
        await cb.answer("Не знайдено", show_alert=True)
        return
    if offer["status"] == "closed":
        await cb.answer("Оголошення вже закрите", show_alert=True)
        return

    db.set_offer_winner(oid, pid)
    db.set_offer_status(oid, "closed")
    db.increment_wins(proposal["user_id"])
    offer = db.get_offer(oid)
    await update_channel_post(offer)

    # Повідомити переможця
    contact_block = ""
    if offer.get("contact_name") or offer.get("contact_phone"):
        contact_block = (
            "\n\n<b>Контакт замовника:</b>\n"
            f"{html_escape(offer.get('contact_name') or '—')}"
            f" · <code>{html_escape(offer.get('contact_phone') or '—')}</code>"
        )
    try:
        await bot.send_message(
            proposal["user_id"],
            f"🏆 <b>Ваша пропозиція прийнята!</b>\n\n"
            f"Оголошення #{oid}: {html_escape(offer['route_from'])} → "
            f"{html_escape(offer['route_to'])}\n"
            f"📦 {html_escape(offer['cargo'])} · "
            f"⚖️ {fmt_weight(offer['weight_t'])} т"
            + contact_block,
        )
    except (TelegramForbiddenError, TelegramBadRequest) as e:
        logger.warning("Не вдалось повідомити переможця: %s", e)

    # Повідомити тих, хто програв
    for p in db.list_proposals(oid):
        if p["id"] == pid:
            continue
        try:
            await bot.send_message(
                p["user_id"],
                f"Дякуємо за пропозицію на оголошення #{oid}. "
                f"Цього разу обрано іншого перевізника. Будемо раді новим заявкам!",
            )
        except (TelegramForbiddenError, TelegramBadRequest):
            pass

    await cb.answer("Переможця обрано", show_alert=True)
    link = user_profile_link(
        proposal["user_id"],
        proposal.get("username"),
        proposal.get("first_name"),
    )
    await cb.message.answer(
        f"🏆 Переможець по #{oid}: {link}\n"
        f"Оголошення закрите. Перевізнику надіслано контакт замовника."
    )


@router.callback_query(F.data.startswith("ban:"))
async def cb_ban(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    _, uid_s, oid_s = cb.data.split(":")
    await state.set_state(BlacklistStates.reason)
    await state.update_data(ban_user_id=int(uid_s), ban_back_offer=int(oid_s))
    await cb.message.answer(
        "Вкажіть <b>причину блокування</b> (буде показана перевізнику). "
        "Можна просто «—» щоб без причини.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text=BTN_CANCEL)]],
            resize_keyboard=True,
        ),
    )
    await cb.answer()


@router.message(BlacklistStates.reason)
async def ban_reason(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    reason = (message.text or "").strip()
    if reason == "—":
        reason = ""
    data = await state.get_data()
    uid = data["ban_user_id"]
    db.set_blacklist(uid, True, reason)
    await state.clear()
    await message.answer(
        f"🚫 Користувача {uid} заблоковано."
        + (f"\nПричина: {html_escape(reason)}" if reason else ""),
        reply_markup=admin_menu_keyboard(message.from_user.id),
    )
    try:
        await bot.send_message(
            uid,
            "🚫 Вас було заблоковано в системі оголошень."
            + (f"\nПричина: {html_escape(reason)}" if reason else ""),
        )
    except (TelegramForbiddenError, TelegramBadRequest):
        pass


@router.callback_query(F.data.startswith("unban:"))
async def cb_unban(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    _, uid_s, _ = cb.data.split(":")
    uid = int(uid_s)
    db.set_blacklist(uid, False)
    await cb.answer("Розблоковано", show_alert=True)
    try:
        await bot.send_message(
            uid, "✅ Вас розблоковано. Можете подавати пропозиції."
        )
    except (TelegramForbiddenError, TelegramBadRequest):
        pass


# ────────────────────  Закрити / відкрити ────────────────────
async def _close_offer(target, offer_id: int, user_id: int):
    if not is_admin(user_id):
        return
    offer = db.get_offer(offer_id)
    if not offer:
        await target.answer("Не знайдено.")
        return
    db.set_offer_status(offer_id, "closed")
    offer = db.get_offer(offer_id)
    await update_channel_post(offer)
    await target.answer(f"🔴 Оголошення #{offer_id} закрите.")


async def _reopen_offer(target, offer_id: int, user_id: int):
    if not is_admin(user_id):
        return
    offer = db.get_offer(offer_id)
    if not offer:
        await target.answer("Не знайдено.")
        return
    new_status = "in_progress" if db.count_proposals(offer_id) else "open"
    db.set_offer_status(offer_id, new_status)
    offer = db.get_offer(offer_id)
    await update_channel_post(offer)
    await target.answer(
        f"{status_emoji(new_status)} Оголошення #{offer_id} знову активне."
    )


@router.message(Command("close"))
async def cmd_close(message: Message, command: CommandObject):
    if not command.args:
        return
    try:
        oid = int(command.args.strip())
    except ValueError:
        return
    await _close_offer(message, oid, message.from_user.id)


@router.message(Command("reopen"))
async def cmd_reopen(message: Message, command: CommandObject):
    if not command.args:
        return
    try:
        oid = int(command.args.strip())
    except ValueError:
        return
    await _reopen_offer(message, oid, message.from_user.id)


@router.callback_query(F.data.startswith("close:"))
async def cb_close(cb: CallbackQuery):
    oid = int(cb.data.split(":", 1)[1])
    await _close_offer(cb.message, oid, cb.from_user.id)
    await cb.answer()


@router.callback_query(F.data.startswith("reopen:"))
async def cb_reopen(cb: CallbackQuery):
    oid = int(cb.data.split(":", 1)[1])
    await _reopen_offer(cb.message, oid, cb.from_user.id)
    await cb.answer()


# ────────────────────  Редагування оголошення  ────────────────────
@router.callback_query(F.data.startswith("edit:"))
async def cb_edit_start(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    oid = int(cb.data.split(":", 1)[1])
    offer = db.get_offer(oid)
    if not offer:
        await cb.answer("Не знайдено", show_alert=True)
        return
    await cb.message.answer(
        f"Редагування оголошення #{oid}. Оберіть поле:",
        reply_markup=edit_field_keyboard(oid),
    )
    await cb.answer()


@router.callback_query(F.data.startswith("editcancel:"))
async def cb_edit_cancel(cb: CallbackQuery):
    try:
        await cb.message.delete()
    except TelegramBadRequest:
        pass
    await cb.answer("Закрито")


FIELD_PROMPTS = {
    "route_from": "Введіть новий <b>Маршрут ВІД</b>:",
    "route_to": "Введіть новий <b>Маршрут ДО</b>:",
    "cargo": "Введіть новий <b>Вантаж</b>:",
    "weight_t": "Введіть нову <b>Вагу (тонн)</b>:",
    "load_date": "Введіть нову <b>Дату завантаження</b>:",
    "extra_info": "Введіть нову <b>Примітку</b>:",
    "contact_name": "Введіть нову <b>Контактну особу</b>:",
    "contact_phone": "Введіть новий <b>Телефон контакту</b>:",
}


@router.callback_query(F.data.startswith("editf:"))
async def cb_edit_field(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    _, oid_s, field = cb.data.split(":")
    await state.set_state(EditOfferStates.new_value)
    await state.update_data(edit_offer_id=int(oid_s), edit_field=field)
    await cb.message.answer(
        FIELD_PROMPTS.get(field, "Введіть нове значення:"),
        reply_markup=new_offer_keyboard(),
    )
    await cb.answer()


@router.message(EditOfferStates.new_value)
async def edit_new_value(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    data = await state.get_data()
    oid = data["edit_offer_id"]
    field = data["edit_field"]
    value: object = message.text.strip()

    if field == "weight_t":
        parsed = parse_tonnage(value)
        if not parsed:
            await message.answer("Некоректна вага. Введіть число 0-100000.")
            return
        value = parsed
    elif field == "contact_phone":
        parsed = parse_phone(value)
        if not parsed:
            await message.answer("Некоректний телефон. Напр. +380501234567.")
            return
        value = parsed

    db.update_offer_fields(oid, **{field: value})
    offer = db.get_offer(oid)
    await update_channel_post(offer)
    await state.clear()
    await message.answer(
        f"✅ Оновлено поле <b>{field}</b> в оголошенні #{oid}.",
        reply_markup=admin_menu_keyboard(message.from_user.id),
    )


# ────────────────────  Адміни (додати / видалити)  ────────────────────
async def _show_admins(target_message: Message, viewer_id: int):
    admins = db.list_admins()
    lines = [
        f"<b>👑 Головний адмін (з env):</b> <code>{SUPER_ADMIN_ID}</code>",
        "",
        "<b>Додаткові адміни:</b>",
    ]
    if admins:
        for a in admins:
            name = a.get("first_name") or a.get("username") or ""
            uname = f"@{a['username']}" if a.get("username") else ""
            lines.append(
                f"• <code>{a['user_id']}</code> "
                f"{html_escape(name)} {uname}"
            )
    else:
        lines.append("—")
    await target_message.answer(
        "\n".join(lines),
        reply_markup=admins_list_inline(admins, is_super_admin(viewer_id)),
    )


@router.message(F.text == ADMIN_BTN_ADMINS)
async def btn_admins(message: Message):
    if not is_super_admin(message.from_user.id):
        return
    await _show_admins(message, message.from_user.id)


# ────────────────────  Модерація реєстрацій  ────────────────────
async def _show_pending_registrations(message: Message):
    pending = db.list_pending_users()
    if not pending:
        await message.answer(
            "Немає заявок на реєстрацію.",
            reply_markup=admin_menu_keyboard(message.from_user.id),
        )
        return
    await message.answer(
        f"<b>📝 Заявки на реєстрацію ({len(pending)}):</b>"
    )
    for u in pending:
        link = user_profile_link(
            u["user_id"], u.get("username"), u.get("first_name")
        )
        text = (
            f"Користувач: {link}\n"
            f"ПІБ / Назва: <b>{html_escape(u.get('full_name'))}</b>\n"
            f"Телефон: <code>{html_escape(u.get('phone'))}</code>\n"
            f"ЄДРПОУ / ІПН: <code>{html_escape(u.get('edrpou') or '—')}</code>\n"
            f"Подано: {html_escape((u.get('submitted_at') or '')[:16])}"
        )
        await message.answer(
            text, reply_markup=_pending_review_keyboard(u["user_id"])
        )


@router.message(F.text.startswith(ADMIN_BTN_PENDING))
async def btn_pending(message: Message):
    if not is_admin(message.from_user.id):
        return
    await _show_pending_registrations(message)


@router.message(Command("pending"))
async def cmd_pending(message: Message):
    if not is_admin(message.from_user.id):
        return
    await _show_pending_registrations(message)


@router.callback_query(F.data.startswith("approve:"))
async def cb_approve(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    uid = int(cb.data.split(":", 1)[1])
    u = db.get_user(uid)
    if not u:
        await cb.answer("Користувача не знайдено", show_alert=True)
        return
    if not u.get("is_pending"):
        await cb.answer(
            "Заявка вже оброблена іншим адміном", show_alert=True
        )
        # оновити повідомлення — прибрати кнопки
        try:
            current_status = ""
            if u.get("is_registered"):
                current_status = "✅ Вже схвалено"
            elif u.get("is_blacklisted"):
                current_status = "❌ Вже відхилено"
            new_text = (cb.message.html_text or "") + f"\n\n<i>{current_status}</i>"
            await cb.message.edit_text(new_text, reply_markup=None)
        except TelegramBadRequest:
            pass
        return

    db.approve_user(uid)

    approver_name = cb.from_user.first_name or f"id{cb.from_user.id}"
    # прибираємо кнопки на повідомленні
    try:
        new_text = (
            (cb.message.html_text or "")
            + f"\n\n✅ <i>Схвалено: {html_escape(approver_name)}</i>"
        )
        await cb.message.edit_text(new_text, reply_markup=None)
    except TelegramBadRequest:
        pass

    # повідомляємо перевізника
    try:
        await bot.send_message(
            uid,
            "✅ <b>Вашу реєстрацію схвалено!</b>\n"
            "Тепер ви можете подавати пропозиції на оголошення.\n\n"
            "Для початку — зайдіть у канал і натисніть кнопку під будь-яким "
            "свіжим оголошенням, або напишіть /start.",
            reply_markup=carrier_menu_keyboard(),
        )
    except (TelegramForbiddenError, TelegramBadRequest):
        pass

    await cb.answer("Схвалено ✅", show_alert=False)


@router.callback_query(F.data.startswith("reject:"))
async def cb_reject(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer()
        return
    uid = int(cb.data.split(":", 1)[1])
    u = db.get_user(uid)
    if not u:
        await cb.answer("Не знайдено", show_alert=True)
        return
    if not u.get("is_pending"):
        await cb.answer("Заявку вже оброблено", show_alert=True)
        return
    await state.set_state(RejectReasonStates.waiting_reason)
    await state.update_data(
        reject_user_id=uid,
        reject_source_chat=cb.message.chat.id,
        reject_source_msg=cb.message.message_id,
    )
    await cb.message.answer(
        "Вкажіть <b>причину відхилення</b> (побачить перевізник). "
        "Можна «—» щоб без причини.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text=BTN_CANCEL)]],
            resize_keyboard=True,
        ),
    )
    await cb.answer()


@router.message(RejectReasonStates.waiting_reason)
async def reject_reason_got(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    reason = (message.text or "").strip()
    if reason == "—":
        reason = ""

    data = await state.get_data()
    uid = data.get("reject_user_id")
    src_chat = data.get("reject_source_chat")
    src_msg = data.get("reject_source_msg")
    await state.clear()

    if not uid:
        await message.answer(
            "Сесія загублена.",
            reply_markup=admin_menu_keyboard(message.from_user.id),
        )
        return

    db.reject_user(uid, reason)
    rejecter_name = message.from_user.first_name or f"id{message.from_user.id}"

    # Прибираємо кнопки з оригінального повідомлення з заявкою
    if src_chat and src_msg:
        try:
            await bot.edit_message_reply_markup(
                chat_id=src_chat,
                message_id=src_msg,
                reply_markup=None,
            )
        except TelegramBadRequest:
            pass

    await message.answer(
        f"❌ Користувача <code>{uid}</code> відхилено."
        + (f"\nПричина: {html_escape(reason)}" if reason else ""),
        reply_markup=admin_menu_keyboard(message.from_user.id),
    )
    try:
        await bot.send_message(
            uid,
            "❌ На жаль, вашу заявку на реєстрацію відхилено."
            + (f"\nПричина: {html_escape(reason)}" if reason else "")
            + "\n\nЯкщо вважаєте це помилкою — зв'яжіться з адміністратором "
            "безпосередньо.",
        )
    except (TelegramForbiddenError, TelegramBadRequest):
        pass


@router.message(Command("admins"))
async def cmd_admins(message: Message):
    if not is_super_admin(message.from_user.id):
        return
    await _show_admins(message, message.from_user.id)


@router.callback_query(F.data == "addadmin")
async def cb_addadmin(cb: CallbackQuery, state: FSMContext):
    if not is_super_admin(cb.from_user.id):
        await cb.answer()
        return
    await state.set_state(AddAdminStates.waiting_for_user)
    await cb.message.answer(
        "Щоб додати адміна:\n"
        "• <b>Перешліть</b> будь-яке повідомлення від цієї людини, АБО\n"
        "• Надішліть її <b>user ID</b> (числом).",
        reply_markup=new_offer_keyboard(),
    )
    await cb.answer()


@router.message(AddAdminStates.waiting_for_user)
async def add_admin_got_input(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)

    user_id = None
    username = ""
    first_name = ""

    if (
        message.forward_origin
        and isinstance(message.forward_origin, MessageOriginUser)
    ):
        u = message.forward_origin.sender_user
        user_id = u.id
        username = u.username or ""
        first_name = u.first_name or ""
    elif message.text:
        try:
            user_id = int(message.text.strip())
        except ValueError:
            await message.answer(
                "Не розпізнав. Перешліть повідомлення або надішліть user ID (число)."
            )
            return

    if not user_id:
        await message.answer("Не розпізнав. Спробуйте ще раз.")
        return

    db.add_admin(user_id, username, first_name)
    await state.clear()
    await message.answer(
        f"✅ Користувача <code>{user_id}</code> додано як адміна.",
        reply_markup=admin_menu_keyboard(message.from_user.id),
    )
    try:
        await bot.send_message(
            user_id,
            "🎉 Вас призначено адміністратором. Напишіть /start щоб побачити меню.",
        )
    except (TelegramForbiddenError, TelegramBadRequest):
        await message.answer(
            "⚠️ Не зміг написати цьому користувачу. Нехай він сам напише /start боту."
        )


@router.callback_query(F.data.startswith("rmadmin:"))
async def cb_rmadmin(cb: CallbackQuery):
    if not is_super_admin(cb.from_user.id):
        await cb.answer()
        return
    uid = int(cb.data.split(":", 1)[1])
    ok = db.remove_admin(uid)
    if ok:
        await cb.answer("Видалено", show_alert=True)
    else:
        await cb.answer("Не знайдено")
    await _show_admins(cb.message, cb.from_user.id)


@router.message(Command("addadmin"))
async def cmd_addadmin(message: Message, command: CommandObject):
    if not is_super_admin(message.from_user.id):
        return
    if not command.args:
        await message.answer("Напр. <code>/addadmin 123456789</code>")
        return
    try:
        uid = int(command.args.strip())
    except ValueError:
        return
    db.add_admin(uid)
    await message.answer(f"✅ {uid} додано.")


@router.message(Command("rmadmin"))
async def cmd_rmadmin(message: Message, command: CommandObject):
    if not is_super_admin(message.from_user.id):
        return
    if not command.args:
        return
    try:
        uid = int(command.args.strip())
    except ValueError:
        return
    db.remove_admin(uid)
    await message.answer(f"🗑 {uid} видалено.")


# ────────────────────  Введення ціни (перевізник)  ────────────────────
async def _ensure_registered_or_prompt(message: Message, state: FSMContext) -> bool:
    if is_admin(message.from_user.id):
        return True
    if db.is_blacklisted(message.from_user.id):
        u = db.get_user(message.from_user.id) or {}
        reason = u.get("blacklist_reason") or ""
        await message.answer(
            "🚫 Ви заблоковані."
            + (f"\nПричина: {html_escape(reason)}" if reason else "")
        )
        return False
    if db.is_pending(message.from_user.id):
        await message.answer(
            "⏳ Ваша заявка на реєстрацію ще на перевірці у адміністратора. "
            "Як тільки її схвалять, ви отримаєте сповіщення й зможете подавати пропозиції."
        )
        return False
    if not db.is_registered(message.from_user.id):
        await start_registration(message, state)
        return False
    return True


@router.message(F.text == BTN_WITH_VAT)
async def price_with_vat_start(message: Message, state: FSMContext):
    if not await _ensure_registered_or_prompt(message, state):
        return
    data = await state.get_data()
    rid = data.get("active_request_id")
    if not rid:
        await message.answer(
            "Спершу відкрийте оголошення через кнопку в каналі."
        )
        return
    await state.set_state(PriceInputStates.with_vat)
    await message.answer(
        "Введіть ціну <b>з ПДВ</b>, грн (лише число, напр. <code>15000</code>):",
        reply_markup=price_input_keyboard(),
    )


@router.message(F.text == BTN_WITHOUT_VAT)
async def price_without_vat_start(message: Message, state: FSMContext):
    if not await _ensure_registered_or_prompt(message, state):
        return
    data = await state.get_data()
    rid = data.get("active_request_id")
    if not rid:
        await message.answer(
            "Спершу відкрийте оголошення через кнопку в каналі."
        )
        return
    await state.set_state(PriceInputStates.without_vat)
    await message.answer(
        "Введіть ціну <b>без ПДВ</b>, грн:",
        reply_markup=price_input_keyboard(),
    )


async def _save_price(
    message: Message,
    state: FSMContext,
    field: str,
):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)

    price = parse_price(message.text or "")
    if price is None:
        await message.answer(
            "Некоректна ціна. Введіть число, напр. <code>15000</code> або "
            "<code>12 500.50</code>."
        )
        return

    data = await state.get_data()
    request_id = data.get("active_request_id")
    offer_id = data.get("active_offer_id")
    if not request_id or not offer_id:
        await message.answer("Сесія загублена. Відкрийте оголошення ще раз.")
        await state.clear()
        return

    if field == "with_vat":
        db.update_proposal_price(request_id, price_with_vat=price)
    else:
        db.update_proposal_price(request_id, price_without_vat=price)

    offer = db.get_offer(offer_id)
    proposal = db.get_proposal(request_id)

    # 🟢 → 🟡 при першій пропозиції
    if offer["status"] == "open":
        db.set_offer_status(offer_id, "in_progress")
        offer = db.get_offer(offer_id)
        await update_channel_post(offer)

    await notify_admins_new_price(offer, proposal, message.from_user.id)

    # Повернутись на картку
    await state.set_state(None)
    await message.answer(
        format_offer_for_carrier(offer, request_id, proposal),
        reply_markup=carrier_card_keyboard(),
    )
    await message.answer("✅ Ціну збережено. Адмін отримав сповіщення.")


@router.message(PriceInputStates.with_vat)
async def price_with_vat_save(message: Message, state: FSMContext):
    await _save_price(message, state, "with_vat")


@router.message(PriceInputStates.without_vat)
async def price_without_vat_save(message: Message, state: FSMContext):
    await _save_price(message, state, "without_vat")


# ────────────────────  «Передати контакт»  ────────────────────
async def _show_contact_request(message: Message):
    await message.answer(
        "Поділіться номером кнопкою нижче. Адмін отримає його миттєво.",
        reply_markup=share_phone_keyboard(),
    )


@router.message(F.text == BTN_CONTACT)
async def btn_contact(message: Message, state: FSMContext):
    if not await _ensure_registered_or_prompt(message, state):
        return
    await _show_contact_request(message)


@router.message(Command("contact"))
async def cmd_contact(message: Message, state: FSMContext):
    if not await _ensure_registered_or_prompt(message, state):
        return
    await _show_contact_request(message)


@router.message(F.contact)
async def got_contact(message: Message, state: FSMContext):
    if is_admin(message.from_user.id):
        return
    phone = parse_phone(message.contact.phone_number)
    if not phone:
        await message.answer("Не вдалось обробити номер.")
        return
    db.update_user_phone(message.from_user.id, phone)
    link = user_profile_link(
        message.from_user.id,
        message.from_user.username or "",
        message.from_user.first_name or "",
    )
    await notify_admins(
        f"📱 Перевізник {link} поділився контактом: "
        f"<code>{html_escape(phone)}</code>"
    )
    data = await state.get_data()
    if data.get("active_request_id") and data.get("active_offer_id"):
        offer = db.get_offer(data["active_offer_id"])
        proposal = db.get_proposal(data["active_request_id"])
        if offer and proposal:
            await message.answer(
                format_offer_for_carrier(offer, proposal["id"], proposal),
                reply_markup=carrier_card_keyboard(),
            )
            return
    await message.answer(
        "✅ Дякуємо, номер передано адміну.",
        reply_markup=carrier_menu_keyboard(),
    )


# ────────────────────  Broadcast (розсилка)  ────────────────────
@router.message(F.text == ADMIN_BTN_BROADCAST)
async def btn_broadcast(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        return
    await state.set_state(BroadcastStates.text)
    await message.answer(
        "Введіть текст розсилки. Всі зареєстровані (не-заблоковані) користувачі "
        "отримають його в приватний чат з ботом. HTML-теги підтримуються.",
        reply_markup=new_offer_keyboard(),
    )


@router.message(BroadcastStates.text)
async def broadcast_text(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    text = message.html_text or message.text or ""
    if len(text.strip()) < 2:
        await message.answer("Текст занадто короткий.")
        return
    await state.update_data(broadcast_text=text)
    await state.set_state(BroadcastStates.confirm)
    users = db.list_all_users_for_broadcast()
    await message.answer(
        f"<b>Попередній перегляд:</b>\n\n{text}\n\n"
        f"Надіслати це <b>{len(users)}</b> користувачам?",
        reply_markup=broadcast_confirm_keyboard(),
    )


@router.message(BroadcastStates.confirm)
async def broadcast_send(message: Message, state: FSMContext):
    if message.text == BTN_CANCEL:
        return await _universal_cancel(message, state)
    if message.text != "📤 Розіслати":
        await message.answer("Натисніть «📤 Розіслати» або «❌ Скасувати».")
        return

    data = await state.get_data()
    text = data["broadcast_text"]
    users = db.list_all_users_for_broadcast()
    await state.clear()

    await message.answer(
        f"Починаю розсилку на {len(users)} користувачів…",
        reply_markup=admin_menu_keyboard(message.from_user.id),
    )

    sent = 0
    failed = 0
    # Telegram ~30 msg/sec глобальний ліміт
    for uid in users:
        try:
            await bot.send_message(uid, text)
            sent += 1
        except (TelegramForbiddenError, TelegramBadRequest) as e:
            failed += 1
            logger.info("broadcast fail %s: %s", uid, e)
        except Exception:
            failed += 1
            logger.exception("broadcast unexpected")
        await asyncio.sleep(0.04)

    db.log_broadcast(message.from_user.id, text, sent, failed)
    await message.answer(
        f"✅ Розсилка завершена.\nНадіслано: <b>{sent}</b>\nПомилок: <b>{failed}</b>"
    )


# ────────────────────  Статистика  ────────────────────
@router.message(F.text == ADMIN_BTN_STATS)
async def btn_stats(message: Message):
    if not is_admin(message.from_user.id):
        return
    s7 = db.stats_summary(7)
    s30 = db.stats_summary(30)

    def block(label, s):
        avg = (
            f"{fmt_price(s['avg_vat'])} грн"
            if s["avg_vat"]
            else "—"
        )
        routes = "\n".join(
            f"  • {html_escape(r['route'])} — {r['c']}"
            for r in s["top_routes"]
        ) or "  —"
        winners = "\n".join(
            f"  • {html_escape(w.get('full_name') or w.get('first_name') or str(w['user_id']))}"
            f" — {w['wins_count']} перемог"
            for w in s["top_winners"]
        ) or "  —"
        return (
            f"<b>{label}</b>\n"
            f"Оголошень: <b>{s['total']}</b>, закрито: <b>{s['closed']}</b>\n"
            f"Середня мін. ціна з ПДВ: <b>{avg}</b>\n"
            f"Топ маршрути:\n{routes}\n"
            f"Топ перевізники:\n{winners}"
        )

    text = (
        f"📊 <b>Статистика</b>\n"
        f"Зареєстрованих перевізників: <b>{s30['registered_carriers']}</b>\n\n"
        + block("За 7 днів", s7)
        + "\n\n"
        + block("За 30 днів", s30)
    )
    await message.answer(text)


# ────────────────────  Планувальники  ────────────────────
async def scheduler_auto_close():
    """Закриває оголошення, у яких минув auto_close_at + шле нагадування."""
    while True:
        try:
            for offer in db.offers_pending_auto_close():
                db.set_offer_status(offer["id"], "closed")
                offer = db.get_offer(offer["id"])
                await update_channel_post(offer)
                await notify_admins(
                    f"⏳ Оголошення #{offer['id']} автоматично закрите "
                    f"(минув термін). Пропозицій: "
                    f"<b>{db.count_proposals(offer['id'])}</b>."
                )
                logger.info("auto-closed offer %s", offer["id"])

            for offer in db.offers_needing_reminder(REMINDER_BEFORE_HOURS):
                users = db.users_without_proposal_for(offer["id"])
                sent = 0
                for u in users:
                    try:
                        await bot.send_message(
                            u["user_id"],
                            f"⏰ Залишилось ~{REMINDER_BEFORE_HOURS} год до "
                            f"закриття оголошення #{offer['id']}: "
                            f"{html_escape(offer['route_from'])} → "
                            f"{html_escape(offer['route_to'])}.\n"
                            f"Не забудьте подати пропозицію!\n\n"
                            f"<a href=\"https://t.me/{BOT_USERNAME}?start=offer_{offer['id']}\">"
                            "Відкрити оголошення</a>",
                        )
                        sent += 1
                    except (TelegramForbiddenError, TelegramBadRequest):
                        pass
                    await asyncio.sleep(0.05)
                db.mark_reminder_sent(offer["id"])
                logger.info(
                    "reminder for offer %s sent to %d users",
                    offer["id"],
                    sent,
                )
        except Exception:
            logger.exception("scheduler_auto_close error")

        await asyncio.sleep(60)


async def scheduler_daily_backup():
    """Раз на добу (о BACKUP_HOUR_KYIV за Києвом ≈ UTC+2/3) шле БД супер-адміну."""
    # Приблизно Kyiv = UTC+3 (літо). На Railway час UTC.
    kyiv_offset_hours = 3
    while True:
        try:
            now_utc = datetime.utcnow()
            target_utc_hour = (BACKUP_HOUR_KYIV - kyiv_offset_hours) % 24
            target = now_utc.replace(
                hour=target_utc_hour, minute=0, second=0, microsecond=0
            )
            if target <= now_utc:
                target += timedelta(days=1)
            sleep_sec = (target - now_utc).total_seconds()
            logger.info(
                "Next backup in %.1f hours", sleep_sec / 3600
            )
            await asyncio.sleep(sleep_sec)

            if not os.path.exists(DB_PATH):
                logger.warning("Backup: DB file not found at %s", DB_PATH)
                continue
            size_kb = os.path.getsize(DB_PATH) / 1024
            await bot.send_document(
                SUPER_ADMIN_ID,
                FSInputFile(DB_PATH),
                caption=(
                    f"🗄 <b>Щоденний бекап БД</b>\n"
                    f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}\n"
                    f"Розмір: {size_kb:.1f} КБ"
                ),
            )
            logger.info("Backup sent to super admin")
        except Exception:
            logger.exception("scheduler_daily_backup error")
            await asyncio.sleep(600)


# ────────────────────  Глобальний error handler  ────────────────────
@dp.errors()
async def on_error(event: ErrorEvent):
    logger.exception("Unhandled bot error", exc_info=event.exception)
    tb = "".join(
        traceback.format_exception(
            type(event.exception), event.exception, event.exception.__traceback__
        )
    )
    snippet = tb[-3500:]
    try:
        await bot.send_message(
            SUPER_ADMIN_ID,
            f"⚠️ <b>Помилка в боті</b>\n<pre>{html_escape(snippet)}</pre>",
        )
    except Exception:
        pass
    return True


# ────────────────────  main  ────────────────────
async def main():
    db.init()
    logger.info("Старт бота…")

    asyncio.create_task(scheduler_auto_close())
    asyncio.create_task(scheduler_daily_backup())

    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
